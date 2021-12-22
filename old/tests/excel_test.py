from openpyxl import Workbook
from datetime import timezone

_STRING_COL_CACHE = {}

def get_column_letter(idx,):
    """Convert a column index into a column letter
    (3 -> 'C')
    """
    try:
        return _STRING_COL_CACHE[idx]
    except KeyError:
        raise ValueError("Invalid column index {0}".format(idx))

class SubReservationExcelReportService:
    __slots__ = ["queryset", "serializer_class"]

    def __init__(self, queryset, serializer_class):
        self.queryset = queryset
        self.serializer_class = serializer_class

    def prepare_content_data(self):
        data = self.serializer_class(self.queryset, many=True).data
        header_mappings = self.serializer_class().Meta.field_mappings
        headers = [k for k, v in sorted(header_mappings.items(), key=lambda x: x[1][1])]
        max_column_length = max([len(x[0]) for x in header_mappings.values()])
        return data, headers, header_mappings, max_column_length

    def create_workbook(self):
        wb = Workbook()
        ws = wb.active
        data, headers, header_mappings, length = self.prepare_content_data()
        ws.append([str(header_mappings.get(h)[0]) for h in headers])
        for column in ws.columns:
            ws.column_dimensions[get_column_letter(column[0].column)].width = length
        for row_data in data:
            row = []
            for header in headers:
                row.append(row_data.get(header) or "-")
            ws.append(row)
        return wb

    def generate_file_name(self):
        timestamp = timezone.localtime(timezone.now()).strftime("%Y%m%d_%H%M%S%f")
        return f"{timestamp}_report.xlsx"

